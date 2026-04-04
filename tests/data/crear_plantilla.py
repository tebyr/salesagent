from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

wb = Workbook()

# ── Colores ──────────────────────────────────────────────────────────────────
HDR_BG   = "1e3a5f"   # Azul oscuro
HDR_FG   = "FFFFFF"   # Blanco
INS_BG   = "f0f0f0"   # Gris claro
INS_FG   = "666666"   # Gris oscuro
ROW_A    = "FFFFFF"   # Blanco
ROW_B    = "eef2ff"   # Azul muy claro
GOLD     = "fff3cd"   # Amarillo suave para instrucciones hoja 00

def hdr_style():
    return (Font(name="Arial", bold=True, color=HDR_FG, size=10),
            PatternFill("solid", fgColor=HDR_BG),
            Alignment(horizontal="center", vertical="center", wrap_text=True))

def ins_style():
    return (Font(name="Arial", italic=True, color=INS_FG, size=9),
            PatternFill("solid", fgColor=INS_BG),
            Alignment(horizontal="left", vertical="center", wrap_text=True))

def data_style(row_idx):
    bg = ROW_A if row_idx % 2 == 0 else ROW_B
    return (Font(name="Arial", size=10),
            PatternFill("solid", fgColor=bg),
            Alignment(horizontal="left", vertical="center", wrap_text=False))

def apply(cell, font, fill, align):
    cell.font   = font
    cell.fill   = fill
    cell.alignment = align

def write_sheet(ws, headers, instructions, rows, col_widths=None):
    """headers: list[str], instructions: list[str], rows: list[list]"""
    # Fila 1 – encabezados
    f, fi, a = hdr_style()
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=c, value=h)
        apply(cell, f, fi, a)

    # Fila 2 – instrucciones
    fi2, fill2, a2 = ins_style()
    for c, ins in enumerate(instructions, 1):
        cell = ws.cell(row=2, column=c, value=ins)
        apply(cell, fi2, fill2, a2)

    # Filas de datos
    for r_idx, row in enumerate(rows, 3):
        fd, filld, ad = data_style(r_idx)
        for c, val in enumerate(row, 1):
            cell = ws.cell(row=r_idx, column=c, value=val)
            apply(cell, fd, filld, ad)

    # Anchos de columna
    if col_widths:
        for col, w in col_widths.items():
            ws.column_dimensions[col].width = w
    else:
        for c in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(c)].width = 22

    # Freeze row 1
    ws.freeze_panes = "A3"
    ws.row_dimensions[1].height = 30
    ws.row_dimensions[2].height = 42


# ═══════════════════════════════════════════════════════════════════════════════
# 00_Instrucciones
# ═══════════════════════════════════════════════════════════════════════════════
ws0 = wb.active
ws0.title = "00_Instrucciones"

instrucciones_generales = [
    ("Este libro contiene las plantillas de datos de prueba para el modelo Sales Agent SaaS.",),
    ("",),
    ("ORDEN DE DILIGENCIAMIENTO:",),
    ("1. 01_Configuracion  – Datos del tenant (distribuidora). Diligenciar primero.",),
    ("2. 02_Vendedores     – Equipo de vendedores. Cada fila es un vendedor.",),
    ("3. 03_Rutas          – Rutas comerciales. Referencian vendedores por nombre.",),
    ("4. 04_Propietarios   – Propietarios de establecimientos (business_owners).",),
    ("5. 05_Clientes       – Establecimientos. Referencian propietarios y rutas.",),
    ("6. 06_Productos      – Catálogo de productos con embalajes y precios.",),
    ("7. 07_Pedidos_Historicos – Pedidos ya realizados. Referencian clientes y vendedores.",),
    ("8. 08_Items_Pedidos  – Líneas de cada pedido. Referencian pedidos y productos.",),
    ("9. 09_Metas_Beneficios – Metas por vendedor y beneficios por cumplimiento.",),
    ("",),
    ("REGLAS GENERALES:",),
    ("• Teléfonos: siempre en formato E.164, ej: +573001234567 (Colombia = +57 + 10 dígitos)",),
    ("• Fechas: formato YYYY-MM-DD, ej: 2026-02-15",),
    ("• Montos: en pesos colombianos (COP) sin puntos ni comas, ej: 8000000",),
    ("• Porcentajes: sin símbolo %, como decimal o entero, ej: 19 (para 19%)",),
    ("• Campos opcionales: dejar en blanco si no aplica. No borrar la columna.",),
    ("• erp_id: código del registro en el ERP del tenant. Dejar vacío si no hay ERP.",),
    ("",),
    ("IMPORTANTE — Unicidad de teléfonos:",),
    ("Un número de teléfono NO puede aparecer a la vez en Vendedores y en Clientes dentro del mismo tenant.",),
    ("El sistema rechazará el registro si detecta este conflicto.",),
]

for r_idx, row in enumerate(instrucciones_generales, 1):
    cell = ws0.cell(row=r_idx, column=1, value=row[0])
    is_title = row[0].startswith(("Este libro", "ORDEN", "REGLAS", "IMPORTANTE"))
    cell.font = Font(name="Arial", bold=is_title, size=11 if is_title else 10,
                     color="1e3a5f" if is_title else "333333")
    cell.alignment = Alignment(wrap_text=True, vertical="top")

ws0.column_dimensions["A"].width = 90
ws0.freeze_panes = None


# ═══════════════════════════════════════════════════════════════════════════════
# 01_Configuracion
# ═══════════════════════════════════════════════════════════════════════════════
ws1 = wb.create_sheet("01_Configuracion")
cfg_headers = ["campo", "valor"]
cfg_ins     = ["Nombre del parámetro (no modificar)", "Valor para este tenant"]
cfg_rows = [
    ["nombre_empresa",          "Distribuciones La Garantía"],
    ["nit",                     "900.123.456-7"],
    ["ciudad",                  "Magangué"],
    ["departamento",            "Bolívar"],
    ["pais",                    "Colombia"],
    ["nombre_agente",           "AsesorBot"],
    ["email_gerencia",          "rodolfo@garantia.co"],
    ["zona_horaria",            "America/Bogota"],
    ["hora_briefing",           "06:30"],
    ["hora_resumen_diario",     "18:30"],
    ["hora_reporte_rendimiento","20:00"],
    ["hora_reporte_gerencia",   "07:00"],
    ["erp_sistema",             ""],
    ["segmentos_clientes",      "Oro, Plata, Bronce"],
]
write_sheet(ws1, cfg_headers, cfg_ins, cfg_rows,
            col_widths={"A": 30, "B": 40})


# ═══════════════════════════════════════════════════════════════════════════════
# 02_Vendedores
# ═══════════════════════════════════════════════════════════════════════════════
ws2 = wb.create_sheet("02_Vendedores")
v_headers = ["nombre_completo","telefono","email","zona","meta_mensual_cop","erp_id","notas"]
v_ins = [
    "Nombre y apellido del vendedor",
    "Formato E.164: +573XXXXXXXXX",
    "Correo corporativo (opcional)",
    "Zona o sector que cubre",
    "Meta de venta mensual en COP sin puntos",
    "ID del vendedor en el ERP (opcional)",
    "Observaciones"
]
v_rows = [
    ["Manuel Iglesias",  "+573172586336", "manuel@garantia.co",  "Zona Norte", 8000000,  "EMP-001", "Ruta MG-001"],
    ["Francia Zuleta",   "+573173715849", "francia@garantia.co", "Zona Sur",   10000000, "EMP-002", "Ruta MG-002"],
]
write_sheet(ws2, v_headers, v_ins, v_rows,
            col_widths={"A":25,"B":20,"C":28,"D":15,"E":20,"F":12,"G":30})


# ═══════════════════════════════════════════════════════════════════════════════
# 03_Rutas
# ═══════════════════════════════════════════════════════════════════════════════
ws3 = wb.create_sheet("03_Rutas")
r_headers = [
    "codigo_ruta","nombre","zona",
    "dias_visita","dias_entrega",
    "horario_lun_vie_inicio","horario_lun_vie_fin",
    "horario_sab_inicio","horario_sab_fin",
    "hora_corte_pedidos",
    "vendedor_asignado","erp_id","notas"
]
r_ins = [
    "Código único de ruta, ej: MG-001",
    "Nombre descriptivo de la ruta",
    "Zona geográfica",
    "Días de visita: 1=Lun,2=Mar,3=Mié,4=Jue,5=Vie,6=Sáb (separados por coma)",
    "Días de entrega (mismo formato)",
    "Hora inicio Lun-Vie HH:MM",
    "Hora fin Lun-Vie HH:MM",
    "Hora inicio Sábado HH:MM",
    "Hora fin Sábado HH:MM",
    "Hora límite toma de pedidos HH:MM",
    "Nombre del vendedor asignado (debe existir en hoja 02)",
    "ID en ERP (opcional)",
    "Observaciones"
]
r_rows = [
    ["MG-001","Ruta Norte Magangué","Zona Norte","1,2,3,4,5,6","3","07:30","16:00","07:30","13:00","15:30","Manuel Iglesias","RUT-001","Entrega miércoles"],
    ["MG-002","Ruta Sur Magangué",  "Zona Sur",  "1,2,3,4,5,6","3","08:00","17:00","08:00","13:00","16:00","Francia Zuleta", "RUT-002","Entrega miércoles"],
]
write_sheet(ws3, r_headers, r_ins, r_rows,
            col_widths={"A":12,"B":28,"C":15,"D":18,"E":16,"F":22,"G":20,"H":20,"I":18,"J":22,"K":22,"L":12,"M":30})


# ═══════════════════════════════════════════════════════════════════════════════
# 04_Propietarios
# ═══════════════════════════════════════════════════════════════════════════════
ws4 = wb.create_sheet("04_Propietarios")
p_headers = ["cedula_nit","nombre_completo","telefono","email","direccion","ciudad","notas"]
p_ins = [
    "CC o NIT del propietario (sin puntos ni guiones)",
    "Nombre completo del propietario",
    "Teléfono personal E.164 (diferente al teléfono del negocio)",
    "Correo del propietario (opcional)",
    "Dirección de residencia (opcional)",
    "Ciudad de residencia",
    "Observaciones"
]
p_rows = [
    ["1065821347","Jesús María Contreras", "+573115540023","",                      "Calle 3 #12-45",          "Magangué","Propietario Panadería El Buen Pan"],
    ["1065734892","Rosa Elena Martínez",   "+573124478901","rosa.m@hotmail.com",     "Carrera 7 #5-18",         "Magangué","Propietaria Tienda La Esquina"],
    ["1065609234","Facundo Cabrales",       "+573138820045","facundo.c@gmail.com",    "Carrera 2 #21-04 La Ensenada","Magangué","Propietario Ferretería Libertador"],
    ["1065518763","Margarita Perdomo",      "+573159930067","margarita.p@gmail.com",  "Calle 10 #8-32",          "Magangué","Propietaria Restaurante El Sabor"],
]
write_sheet(ws4, p_headers, p_ins, p_rows,
            col_widths={"A":16,"B":28,"C":20,"D":28,"E":35,"F":16,"G":36})


# ═══════════════════════════════════════════════════════════════════════════════
# 05_Clientes
# ═══════════════════════════════════════════════════════════════════════════════
ws5 = wb.create_sheet("05_Clientes")
c_headers = [
    "nombre_negocio","cedula_nit_propietario","telefono_establecimiento","email",
    "tipologia","clasificacion","direccion","barrio","municipio",
    "codigo_ruta","whatsapp_opt_in","limite_credito_cop","plazo_pago_dias","erp_id","notas"
]
c_ins = [
    "Nombre comercial del establecimiento",
    "CC/NIT del propietario — debe existir en hoja 04",
    "Teléfono WA del negocio E.164 (diferente al del propietario y a los vendedores)",
    "Correo del negocio (opcional)",
    "Tienda / Panadería / Ferretería / Restaurante / Minimercado / Supermercado / Droguería / Otro",
    "Oro / Plata / Bronce",
    "Dirección del establecimiento",
    "Barrio",
    "Municipio",
    "Código de ruta — debe existir en hoja 03",
    "SI / NO",
    "Cupo de crédito en COP (0 = contado)",
    "Días de plazo (0 = contado)",
    "ID en ERP (opcional)",
    "Observaciones"
]
c_rows = [
    ["Panadería El Buen Pan", "1065821347","+573201100011","",                  "Panadería",  "Plata", "Calle 4 #15-22",            "Centro",      "Magangué","MG-001","SI",500000, 0,  "CLI-001","Cliente activo, compra semanal"],
    ["Tienda La Esquina",     "1065734892","+573202200022","",                  "Tienda",     "Bronce","Carrera 9 #3-45",           "Barrio Nuevo", "Magangué","MG-001","SI",300000, 0,  "CLI-002","Compra quincenal"],
    ["Ferretería Libertador", "1065609234","+573203300033","ferlib@hotmail.com","Ferretería", "Plata", "Carrera 2 #21-04",          "La Ensenada",  "Magangué","MG-002","SI",1000000,30, "CLI-003","Crédito a 30 días aprobado"],
    ["Restaurante El Sabor",  "1065518763","+573204400044","elsabor@gmail.com", "Restaurante","Bronce","Calle 10 #6-15 Local 2",    "El Progreso",  "Magangué","MG-002","SI",0,      0,  "CLI-004","Solo contado"],
]
write_sheet(ws5, c_headers, c_ins, c_rows,
            col_widths={"A":28,"B":22,"C":22,"D":24,"E":16,"F":12,"G":32,"H":16,"I":14,"J":12,"K":14,"L":18,"M":16,"N":12,"O":30})


# ═══════════════════════════════════════════════════════════════════════════════
# 06_Productos
# ═══════════════════════════════════════════════════════════════════════════════
ws6 = wb.create_sheet("06_Productos")
pr_headers = [
    "erp_sku","nombre_producto","proveedor","marca","categoria","subcategoria",
    "es_obligatorio","tipo_iva","tasa_iva_pct",
    "unidad_base","precio_base_cop",
    "nombre_embalaje_1","unidades_embalaje_1","precio_embalaje_1_cop",
    "nombre_embalaje_2","unidades_embalaje_2","precio_embalaje_2_cop",
    "erp_id","notas"
]
pr_ins = [
    "Código SKU en el ERP",
    "Nombre comercial del producto",
    "Nombre del proveedor/fabricante",
    "Marca comercial",
    "Categoría (ej: Aceites, Granos, Aseo, Bebidas)",
    "Subcategoría (ej: Aceite de Girasol, Arroz Blanco)",
    "SI si es producto 'imperdonable' / NO si es opcional",
    "iva / ipoconsumo / exento",
    "Tasa en % (ej: 19, 5, 0) — canasta familiar = 0",
    "Nombre de la unidad mínima de venta (Unidad, Kilo, Litro)",
    "Precio de lista de la unidad base en COP",
    "Nombre del primer embalaje (ej: Caja x12)",
    "Cuántas unidades base trae",
    "Precio del embalaje en COP",
    "Nombre del segundo embalaje (dejar vacío si no aplica)",
    "Unidades base del segundo embalaje",
    "Precio del segundo embalaje en COP",
    "ID en ERP (opcional)",
    "Observaciones"
]
pr_rows = [
    # SKU, Nombre, Proveedor, Marca, Categoria, Subcateg, Oblig, IVA, Tasa, UBase, PBase, Emb1, Uds1, P1, Emb2, Uds2, P2, ERP, Notas
    ["ACE-001","Aceite de Girasol 3L",   "Aceites del Campo","Girasol Premier","Aceites","Aceite de Girasol","SI","exento",0,  "Unidad",18500,"Caja x6",  6,111000,"","","","","Canasta familiar, IVA 0%"],
    ["ACE-002","Aceite de Palma 1L",     "Aceites del Campo","Girasol Premier","Aceites","Aceite de Palma",  "NO","exento",0,  "Unidad",8200, "Caja x12",12, 98400,"","","","","Canasta familiar, IVA 0%"],
    ["ARR-001","Arroz Diana 5kg",        "Molinos Roa",      "Diana",          "Granos", "Arroz Blanco",     "SI","exento",0,  "Bulto", 22000,"Bulto x5",  5,110000,"Palet x50",50,1100000,"","Presentación 5kg más vendida"],
    ["ARR-002","Arroz Supremo 1kg",      "Molinos Roa",      "Supremo",        "Granos", "Arroz Blanco",     "NO","exento",0,  "Unidad",4800, "Caja x20",20, 96000,"","","","","Presentación unitaria"],
    ["AZU-001","Azúcar Blanca Riopaila 2kg","Riopaila Castilla","Riopaila",   "Azúcar",  "Azúcar Blanca",    "SI","exento",0,  "Unidad",7500, "Caja x10",10, 75000,"","","","","Canasta familiar, IVA 0%"],
    ["SAL-001","Sal Refisal 1kg",         "Refisal",         "Refisal",        "Sal",    "Sal de Mesa",      "SI","exento",0,  "Unidad",2100, "Caja x24",24, 50400,"","","","","Canasta familiar, IVA 0%"],
    ["PAS-001","Pasta Doria Espagueti 500g","Grupo Nutresa",  "Doria",          "Pastas", "Espagueti",        "SI","exento",0,  "Unidad",3200, "Caja x24",24, 76800,"","","","","Canasta familiar, IVA 0%"],
    ["JAB-001","Jabón Rey x3 und",        "Unilever",        "Rey",            "Aseo",   "Jabón de Lavar",   "NO","iva",   19, "Paquete",8900,"Caja x12",12,106800,"","","","","IVA 19% aseo"],
    ["DET-001","Detergente Ariel 1kg",    "P&G",             "Ariel",          "Aseo",   "Detergente Polvo", "NO","iva",   19, "Bolsa", 14500,"Caja x9",   9,130500,"","","","","IVA 19% aseo"],
    ["CAF-001","Café Colcafé Granulado 150g","Grupo Nutresa","Colcafé",        "Bebidas","Café Soluble",     "SI","ipoconsumo",8,"Frasco",9500,"Caja x12",12,114000,"","","","","INC 8% bebidas con cafeína"],
    ["PAN-001","Panela Redonda 1kg",      "Trapiche El Palmar","El Palmar",     "Panela", "Panela Molida",    "SI","exento",0,  "Unidad",4200, "Caja x12",12, 50400,"Bulto x30",30,126000,"","Canasta familiar, IVA 0%"],
    ["LEN-001","Lenteja Verde 500g",      "Granos del Llano","Cosecha",        "Granos", "Legumbres",        "NO","exento",0,  "Unidad",5500, "Caja x20",20,110000,"","","","","Canasta familiar, IVA 0%"],
]
write_sheet(ws6, pr_headers, pr_ins, pr_rows,
            col_widths={"A":12,"B":30,"C":22,"D":18,"E":14,"F":20,"G":13,"H":14,"I":12,
                        "J":12,"K":16,"L":14,"M":16,"N":18,"O":14,"P":16,"Q":18,"R":12,"S":30})


# ═══════════════════════════════════════════════════════════════════════════════
# 07_Pedidos_Historicos
# ═══════════════════════════════════════════════════════════════════════════════
ws7 = wb.create_sheet("07_Pedidos_Historicos")
ped_headers = [
    "referencia_pedido","nombre_cliente","nombre_vendedor","fecha_pedido","fecha_entrega",
    "estado","erp_id_factura",
    "total_pedido_cop","total_facturado_cop","total_devoluciones_cop","total_neto_cop",
    "canal_origen","notas"
]
ped_ins = [
    "Código único del pedido en este SaaS, ej: PED-2026-001",
    "Nombre exacto del cliente — debe existir en hoja 05",
    "Nombre exacto del vendedor — debe existir en hoja 02",
    "Fecha toma de pedido YYYY-MM-DD",
    "Fecha entrega real YYYY-MM-DD (vacío si no entregado)",
    "pedido / facturado / facturado_devolucion_parcial / cancelado",
    "Número de factura en el ERP (vacío si no facturado)",
    "Valor total del pedido original en COP",
    "Valor facturado por el ERP (puede diferir del pedido)",
    "Valor de devoluciones en COP (0 si no hubo)",
    "Neto final = facturado - devoluciones (fórmula: =H-I)",
    "vendedor / agente_wa / admin",
    "Observaciones"
]
ped_rows = [
    ["PED-2026-001","Panadería El Buen Pan", "Manuel Iglesias","2026-02-05","2026-02-07","facturado",          "FE-00100001",96700, 96700, 0,     96700,"vendedor",  "Pedido regular quincenal"],
    ["PED-2026-002","Tienda La Esquina",      "Manuel Iglesias","2026-02-10","2026-02-12","facturado",          "FE-00100002",55700, 55700, 0,     55700,"vendedor",  ""],
    ["PED-2026-003","Ferretería Libertador",  "Francia Zuleta", "2026-02-12","2026-02-14","facturado_devolucion_parcial","FE-00100003",134200,134200,18500,115700,"vendedor","Devolvió 1 caja aceite dañada"],
    ["PED-2026-004","Restaurante El Sabor",   "Francia Zuleta", "2026-02-18","2026-02-19","facturado",          "FE-00100004",71300, 71300, 0,     71300,"vendedor",  ""],
    ["PED-2026-005","Panadería El Buen Pan",  "Manuel Iglesias","2026-03-05","2026-03-07","facturado",          "FE-00100005",118900,118900,0,     118900,"vendedor",  "Pedido más grande por temporada"],
    ["PED-2026-006","Tienda La Esquina",      "Manuel Iglesias","2026-03-15","",          "pedido",             "",           48300, 0,     0,     0,     "agente_wa", "Pedido tomado por WA, pendiente ERP"],
]
write_sheet(ws7, ped_headers, ped_ins, ped_rows,
            col_widths={"A":18,"B":26,"C":20,"D":14,"E":14,"F":28,"G":16,
                        "H":18,"I":18,"J":18,"K":16,"L":14,"M":30})


# ═══════════════════════════════════════════════════════════════════════════════
# 08_Items_Pedidos
# ═══════════════════════════════════════════════════════════════════════════════
ws8 = wb.create_sheet("08_Items_Pedidos")
it_headers = [
    "referencia_pedido","erp_sku_producto","embalaje",
    "cant_pedida","precio_unitario_cop","descuento_pct","total_linea_pedido_cop",
    "cant_facturada","precio_facturado_cop","total_linea_factura_cop",
    "cant_devuelta","valor_devolucion_cop","total_neto_linea_cop","notas"
]
it_ins = [
    "FK a hoja 07 — referencia del pedido",
    "FK a hoja 06 — SKU del producto",
    "Unidad / Caja x12 / Bulto / Paquete — tipo de embalaje vendido",
    "Cantidad vendida en unidades del embalaje",
    "Precio unitario del embalaje en COP",
    "Descuento aplicado en % (0 si no hubo)",
    "Total línea pedido = cant × precio × (1 - desc%)",
    "Cantidad facturada (puede ser menor por quiebre de stock)",
    "Precio facturado (puede diferir por ajuste en bodega)",
    "Total línea factura",
    "Cantidad devuelta (0 si no hubo devolución)",
    "Valor devuelto en COP",
    "Neto línea = total factura - devolución",
    "Observaciones"
]
it_rows = [
    # PED-2026-001 — Panadería El Buen Pan
    ["PED-2026-001","ACE-001","Unidad",  2,18500,0,37000, 2,18500,37000,0,0,37000,""],
    ["PED-2026-001","ARR-001","Bulto",   2,22000,0,44000, 2,22000,44000,0,0,44000,""],
    ["PED-2026-001","SAL-001","Caja x24",1, 2100,0, 2100, 1, 2100, 2100,0,0, 2100,"Por unidades sueltas"],
    ["PED-2026-001","PAN-001","Caja x12",1, 4200,0, 4200, 1, 4200, 4200,0,0, 4200,""],
    # PED-2026-002 — Tienda La Esquina
    ["PED-2026-002","JAB-001","Paquete", 2, 8900,0,17800, 2, 8900,17800,0,0,17800,""],
    ["PED-2026-002","PAS-001","Unidad",  6, 3200,0,19200, 6, 3200,19200,0,0,19200,""],
    ["PED-2026-002","AZU-001","Unidad",  2, 7500,0,15000, 2, 7500,15000,0,0,15000,""],
    ["PED-2026-002","CAF-001","Frasco",  1, 9500,0, 9500, 1, 9500, 9500,0,0, 9500,""],
    # PED-2026-003 — Ferretería Libertador (con devolución)
    ["PED-2026-003","ACE-001","Caja x6", 2,111000,0,222000,2,111000,222000,1,111000,111000,"Devolvió 1 caja — producto dañado"],
    ["PED-2026-003","DET-001","Caja x9", 1,130500,0,130500,1,130500,130500,0,0,130500,""],
    ["PED-2026-003","SAL-001","Caja x24",1, 2100,0,  2100, 1,  2100,  2100,0,0,  2100,""],
    # PED-2026-004 — Restaurante El Sabor
    ["PED-2026-004","ACE-002","Caja x12",1,98400, 0,98400, 1,98400, 98400,0,0, 98400,"Aceite de palma para freidoras"],
    ["PED-2026-004","SAL-001","Caja x24",2, 2100, 0, 4200, 2, 2100,  4200,0,0,  4200,""],
    # PED-2026-005 — Panadería El Buen Pan (temporada)
    ["PED-2026-005","ACE-001","Caja x6", 3,111000,0,333000,3,111000,333000,0,0,333000,"Temporada alta"],
    ["PED-2026-005","ARR-001","Bulto x5",1,110000,0,110000,1,110000,110000,0,0,110000,""],
    ["PED-2026-005","AZU-001","Caja x10",1, 75000,0, 75000,1, 75000, 75000,0,0, 75000,""],
    # PED-2026-006 — Tienda La Esquina (pedido WA, no facturado aún)
    ["PED-2026-006","PAS-001","Unidad",  6, 3200,0,19200, 0,0,      0,    0,0,      0,"Pendiente facturación"],
    ["PED-2026-006","JAB-001","Paquete", 1, 8900,0, 8900, 0,0,      0,    0,0,      0,"Pendiente facturación"],
    ["PED-2026-006","CAF-001","Frasco",  2, 9500,0,19000, 0,0,      0,    0,0,      0,"Pendiente facturación"],
    ["PED-2026-006","AZU-001","Unidad",  1, 7500,0, 7500, 0,0,      0,    0,0,      0,"Pendiente facturación"],
]
write_sheet(ws8, it_headers, it_ins, it_rows,
            col_widths={"A":18,"B":14,"C":14,"D":12,"E":20,"F":14,"G":22,
                        "H":16,"I":20,"J":22,"K":14,"L":20,"M":20,"N":28})


# ═══════════════════════════════════════════════════════════════════════════════
# 09_Metas_Beneficios
# ═══════════════════════════════════════════════════════════════════════════════
ws9 = wb.create_sheet("09_Metas_Beneficios")
m_headers = [
    "nombre_vendedor","periodo",
    "meta_ventas_cop","meta_visitas","meta_clientes_nuevos",
    "umbral_bronce_pct","beneficio_bronce",
    "umbral_plata_pct","beneficio_plata",
    "umbral_oro_pct","beneficio_oro","notas"
]
m_ins = [
    "Nombre exacto — debe existir en hoja 02",
    "Período YYYY-MM, ej: 2026-04",
    "Meta de ventas netas en COP",
    "Número de visitas efectivas requeridas",
    "Número de clientes nuevos requeridos",
    "% mínimo de cumplimiento para nivel Bronce (ej: 80)",
    "Descripción del beneficio Bronce (bono en COP o descripción)",
    "% mínimo para nivel Plata (ej: 90)",
    "Descripción beneficio Plata",
    "% mínimo para nivel Oro (ej: 100)",
    "Descripción beneficio Oro",
    "Observaciones"
]
m_rows = [
    ["Manuel Iglesias","2026-04",8000000,20,1,80,"Bono $200.000",90,"Bono $400.000",100,"Bono $700.000 + Premio kit hogar","Meta mensual estándar"],
    ["Francia Zuleta", "2026-04",10000000,22,2,80,"Bono $250.000",90,"Bono $500.000",100,"Bono $900.000 + Premio kit hogar","Meta mensual estándar"],
]
write_sheet(ws9, m_headers, m_ins, m_rows,
            col_widths={"A":22,"B":12,"C":18,"D":14,"E":20,
                        "F":18,"G":28,"H":16,"I":28,"J":14,"K":36,"L":28})


# ── Guardar ───────────────────────────────────────────────────────────────────
output_path = "/Users/oscarmauriciogomezacevedo/claudecode/salesagent/tests/data/plantilla_datos_prueba.xlsx"
wb.save(output_path)
print(f"OK: {output_path}")
